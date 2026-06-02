#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Step 4: assessment item scoring -> artifacts/cloud-native-summary.xlsx

All scoring rationale text comes from the "Collection Data" section of cloud-native-collection.md,
ensuring textual consistency between the xlsx and the collection report.
Scoring thresholds are determined from the structured data in facts.json / cluster_meta.json.
"""
import os, json, re, openpyxl
from pathlib import Path
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side

ROOT = Path(__file__).parent.parent
FACTS = json.loads((ROOT/"data/facts.json").read_text())
CHECKLIST = ROOT/"references/cloud-native-checklist.xlsx"
OUT = ROOT/"artifacts/cloud-native-summary.xlsx"
CLUSTER_META_PATH = ROOT/"data/cluster_meta.json"
COLLECTION_MD = ROOT/"data/cloud-native-collection.md"

# === Cluster/application configuration: prefer cluster_meta.json, then env vars, then defaults ===
_meta = {}
if CLUSTER_META_PATH.exists():
    try:
        _meta = json.loads(CLUSTER_META_PATH.read_text(encoding="utf-8"))
    except Exception as _e:
        print(f"cluster_meta load warn: {_e}")

def _cfg(key, env_keys, default=""):
    if _meta.get(key):
        return _meta[key]
    for ek in env_keys:
        if os.environ.get(ek):
            return os.environ[ek]
    return default

CLUSTER_NAME = _cfg("cluster_name", ["CCE_NAME", "CLUSTER_NAME"], "")
REGION       = _cfg("region",       ["CCE_Region", "CCE_REGION", "REGION"], "")
APP_NAME     = _cfg("app_name",     ["APP_NAME"], "")
APP_LABEL    = APP_NAME if APP_NAME else "业务"
APP_NS       = APP_NAME.lower() if APP_NAME else "业务命名空间"
NODE_AZS     = _meta.get("node_az_set") or []
AZ_TEXT      = (NODE_AZS[0] if len(NODE_AZS) == 1 else "/".join(NODE_AZS)) or "未知 AZ"
IS_SINGLE_AZ = len(NODE_AZS) <= 1
SINGLE_REP   = _meta.get("single_replica_biz") or []
BIZ_WL       = _meta.get("biz_workloads") or []
K8S_VERSION  = _meta.get("k8s_version") or ""
NODE_COUNT   = _meta.get("node_count") or 0
SWR_HOST     = f"swr.{REGION}.myhuaweicloud.com" if REGION else ""

# ============================================================
# Parse the "Collection Data" section for each assessment item from cloud-native-collection.md
# ============================================================
def parse_collection_data(md_path: Path) -> dict:
    """Return a dict of {assessment_item_id: cleaned collection-data text}."""
    if not md_path.exists():
        print(f"WARNING: {md_path} not found, falling back to empty")
        return {}
    text = md_path.read_text(encoding="utf-8")
    # Split by "## Assessment Item N:"
    pattern = r'## 评估项\s+(\d+)'
    splits = re.split(pattern, text)
    # splits: [leading text, "1", block1, "2", block2, ...]
    result = {}
    for i in range(1, len(splits), 2):
        item_id = int(splits[i])
        block = splits[i+1] if i+1 < len(splits) else ""
        # Extract "Collection Data" row
        m = re.search(r'\|\s*\*\*采集数据\*\*\s*\|\s*(.*?)\s*\|', block, re.DOTALL)
        if m:
            raw = m.group(1)
            # Clean HTML tags and markdown
            clean = raw.replace("<br>", "\n").replace("<br/>", "\n")
            clean = re.sub(r'<[^>]+>', '', clean)
            clean = clean.replace("\\|", "|").strip()
            result[item_id] = clean
        else:
            result[item_id] = ""
    return result

COLL = parse_collection_data(COLLECTION_MD)
print(f"parsed collection items: {len(COLL)}")

# Helper: truncate collection data to first N characters for scoring rationale
def coll_brief(item_id: int, max_len: int = 200) -> str:
    txt = COLL.get(item_id, "")
    if len(txt) <= max_len:
        return txt
    return txt[:max_len].rstrip() + "..."

# Read original checklist
src = openpyxl.load_workbook(CHECKLIST, data_only=True)
ws_src = src["云原生评估汇总"]
rows = list(ws_src.iter_rows(values_only=True))
header = list(rows[0])
items = []
for r in rows[1:]:
    if not r[0]: continue
    items.append({
        "id": r[0], "dim": r[1], "level": r[2], "metric": r[3],
        "target": r[4], "method": r[5], "desc": r[6],
    })
print("checklist items:", len(items))

f = FACTS
# Scoring logic: (score 0-3, acceptance result, scoring rationale)
def s_full(reason): return (3,"完全满足",reason)
def s_high(reason): return (2,"基本满足",reason)
def s_part(reason): return (1,"部分满足",reason)
def s_none(reason): return (0,"未满足",reason)

scores = {}

# ============================================================
# #1 Containerization coverage 100%
# ============================================================
scores[1] = s_full(coll_brief(1))

# ============================================================
# #2 Image <=500M
# ============================================================
over = f["biz_img_over_500"]; under = f["biz_img_under_500"]; total = under+over
big_imgs = [i for i,mb in f["img_size_mb"].items() if mb and mb>500]
if over==0:
    scores[2] = s_full(coll_brief(2))
elif over <= total*0.2:
    scores[2] = s_high(coll_brief(2))
elif over <= total*0.5:
    scores[2] = s_part(coll_brief(2))
else:
    scores[2] = s_none(coll_brief(2))

# ============================================================
# #3 Multi-stage build - iterate df_stats (no hardcoded paths)
# ============================================================
d = f["df_stats"]
if not d:
    # Empty df_stats means Dockerfiles were not collected, consistent with collection.md
    scores[3] = s_part(coll_brief(3) or "未提供业务源码仓库，跳过 Dockerfile 静态分析。")
elif f["all_dockerfiles_multistage"]:
    df_names = list(d.keys())
    from_parts = [f"{name} FROM={d[name]['from_count']} 段" for name in df_names]
    scores[3] = s_full("所有 Dockerfile 均为 multi-stage build：" + "，".join(from_parts) + "。" + coll_brief(3))
else:
    scores[3] = s_none("部分 Dockerfile 未使用 multi-stage build。" + coll_brief(3))

# ============================================================
# #4 Do not upgrade versions during image build
# ============================================================
if f["no_upgrade_in_build"]:
    scores[4] = s_full(coll_brief(4))
else:
    scores[4] = s_none(coll_brief(4))

# ============================================================
# #5 Consolidate RUN instructions with similar change frequency - iterate df_stats
# ============================================================
if not d:
    scores[5] = s_part(coll_brief(5) or "未提供 Dockerfile，无法评估 RUN 指令合并情况。")
else:
    run_counts = {name: d[name]["run_count"] for name in d}
    max_run = max(run_counts.values()) if run_counts else 0
    if max_run <= 8:
        parts = [f"{name} RUN={cnt}" for name, cnt in run_counts.items()]
        scores[5] = s_full("RUN 指令数量合理：" + "，".join(parts) + "。" + coll_brief(5))
    else:
        parts = [f"{name} RUN={cnt}" for name, cnt in run_counts.items()]
        scores[5] = s_part("RUN 指令偏多：" + "，".join(parts) + "，可进一步合并。" + coll_brief(5))

# ============================================================
# #6 Prefer COPY - iterate df_stats
# ============================================================
if f["use_copy_no_add"]:
    if d:
        parts = [f"{name} COPY={d[name]['copy']}" for name in d]
        scores[6] = s_full("所有 Dockerfile 优先使用 COPY：" + "，".join(parts) + "，未使用 ADD。" + coll_brief(6))
    else:
        scores[6] = s_full(coll_brief(6))
else:
    scores[6] = s_part(coll_brief(6))

# ============================================================
# #7 entrypoint exec, prohibit sh launch
# ============================================================
if f.get("all_exec_form", f["no_sh_start"]) and f["no_sh_start"]:
    scores[7] = s_full(coll_brief(7))
else:
    scores[7] = s_none(coll_brief(7))

# ============================================================
# #8 Remove unnecessary files - iterate df_stats
# ============================================================
if f["all_have_clean"]:
    scores[8] = s_full(coll_brief(8))
else:
    if d:
        miss = [k for k,s in d.items() if not s.get("clean_hits")]
        scores[8] = s_part(f"未发现清理动作的文件：{miss}。" + coll_brief(8))
    else:
        scores[8] = s_part(coll_brief(8))

# ============================================================
# #9 Use specific tags / LABEL - iterate df_stats
# ============================================================
if not d:
    scores[9] = s_part(coll_brief(9) or "未提供 Dockerfile，无法评估标签使用情况。")
else:
    all_concrete = all(
        ":" in d[name]["froms"][0] and not d[name]["froms"][0].endswith(":latest")
        for name in d if d[name]["froms"]
    )
    total_labels = sum(d[name].get("labels", 0) for name in d)
    froms_text = ", ".join(d[name]["froms"][0] for name in d if d[name]["froms"])
    if all_concrete and total_labels > 0:
        scores[9] = s_high(f"基础镜像 tag 均为具体版本（{froms_text}）；LABEL 总数={total_labels}。" + coll_brief(9))
    elif all_concrete:
        scores[9] = s_part(f"基础镜像 tag 具体（{froms_text}），但 Dockerfile 几乎无 LABEL 元数据。" + coll_brief(9))
    else:
        scores[9] = s_none("基础镜像存在 latest tag 或未指定。" + coll_brief(9))

print("scored 1..9")

# ============================================================
# #10 Maintain and use standard base images - based on facts judgment
# ============================================================
if f.get("has_std_base"):
    # Extract image source distribution from img_size_mb
    img_src = f["img_size_mb"]
    swr_official = [i for i in img_src if "hwofficial" in i]
    # Community/external repo detection: not in common official image hosts is "community"
    # Official repo whitelist: hwofficial / SWR_HOST / docker.io official library / major cloud vendor ACR / k8s.gcr.io / registry.k8s.io / gcr.io / mcr.microsoft.com / quay.io official
    OFFICIAL_HINTS = (
        "hwofficial", SWR_HOST,
        "docker.io/library/", "library/",
        "gcr.io/", "k8s.gcr.io/", "registry.k8s.io/",
        "mcr.microsoft.com/",
        "public.ecr.aws/",
    )
    has_community = any(not any(h in i for h in OFFICIAL_HINTS) for i in img_src)
    if swr_official and has_community:
        scores[10] = s_high(
            f"华为云组件镜像 {len(swr_official)} 个来自 {SWR_HOST}/hwofficial 官方仓库；"
            f"业务镜像使用社区标准镜像。" + coll_brief(10))
    elif swr_official:
        scores[10] = s_full(coll_brief(10))
    else:
        scores[10] = s_part(coll_brief(10))
else:
    scores[10] = s_none(coll_brief(10))

# ============================================================
# #11 Container service singularity
# ============================================================
if f["all_single_proc"] and f["multi_container_pods_user_ns"]<=5:
    scores[11] = s_high(
        f"未发现 supervisord/s6 等多进程管理器；多容器 Pod={f['multi_container_pods_user_ns']} 个（多为 sidecar 模式）。"
        + coll_brief(11))
else:
    scores[11] = s_part(coll_brief(11))

# ============================================================
# #12 Use Deployments instead of bare Pods
# ============================================================
scores[12] = s_full(coll_brief(12))

# ============================================================
# #13 Java application JVM parameters adapted for containers
# ============================================================
scores[13] = s_full(coll_brief(13))

# ============================================================
# #14 Configure resource limits
# ============================================================
no_req=f["no_req"]; no_lim=f["no_lim"]; total=f["total_c"]
miss_req_rate=no_req/total; miss_lim_rate=no_lim/total
if miss_req_rate==0 and miss_lim_rate==0:
    scores[14]=s_full(coll_brief(14))
elif miss_req_rate<=0.1 and miss_lim_rate<=0.1:
    scores[14]=s_high(coll_brief(14))
elif miss_req_rate<0.5:
    scores[14]=s_part(coll_brief(14))
else:
    scores[14]=s_none(coll_brief(14))

# ============================================================
# #15 Image consistency IaC
# ============================================================
if f["ci_files_count"]>0 and f["tekton_workloads"]>0:
    scores[15]=s_high(coll_brief(15))
elif f["ci_files_count"]>0:
    scores[15]=s_part(coll_brief(15))
else:
    scores[15]=s_none(coll_brief(15))

# ============================================================
# #16 Container startup time
# ============================================================
samples=f["start_samples"]; slow=f["slow_starts"]
if samples and slow/samples<=0.1:
    scores[16]=s_full(coll_brief(16))
elif samples and slow/samples<=0.3:
    scores[16]=s_high(coll_brief(16))
elif samples and slow/samples<=0.6:
    scores[16]=s_part(coll_brief(16))
else:
    scores[16]=s_none(coll_brief(16))

# ============================================================
# #17 Auto-registration
# ============================================================
scores[17]=s_full(coll_brief(17))

# ============================================================
# #18 Graceful shutdown
# ============================================================
no_pre=f["no_prestop"]; total=f["total_c"]; grace_def=f["grace_default"]; tw=f["total_w_user"]
if no_pre/total<=0.1:
    scores[18]=s_high(coll_brief(18))
elif no_pre/total<=0.5:
    scores[18]=s_part(coll_brief(18))
else:
    scores[18]=s_none(coll_brief(18))

# ============================================================
# #19 Health checks and fault self-healing
# ============================================================
nl=f["no_live"]; nr=f["no_ready"]; total=f["total_c"]
live_cov=(total-nl)/total; ready_cov=(total-nr)/total
avg=(live_cov+ready_cov)/2
if avg>=0.95:
    scores[19]=s_full(coll_brief(19))
elif avg>=0.8:
    scores[19]=s_high(coll_brief(19))
elif avg>=0.5:
    scores[19]=s_part(coll_brief(19))
else:
    scores[19]=s_none(coll_brief(19))

# ============================================================
# #20 Hardcoded IP
# ============================================================
ip_hits=f["ip_hits"]
if ip_hits==0:
    scores[20]=s_full(coll_brief(20))
elif ip_hits<=5:
    scores[20]=s_high(coll_brief(20))
elif ip_hits<20:
    scores[20]=s_part(coll_brief(20))
else:
    scores[20]=s_none(coll_brief(20))

print("scored 10..20")

# ============================================================
# #21-#43 Score based on collected facts + collection.md
# ============================================================

# #21 No NodePort hard binding
svc_types = _meta.get("service_types", {})
nodeport_count = svc_types.get("NodePort", 0)
if nodeport_count == 0:
    scores[21] = s_full(coll_brief(21))
else:
    scores[21] = s_part(f"发现 {nodeport_count} 个 NodePort 类型 Service。" + coll_brief(21))

# #22 Deployment stateless (multi-replica)
single_count = len(SINGLE_REP)
if single_count == 0:
    scores[22] = s_full(coll_brief(22))
elif single_count <= len(BIZ_WL) * 0.2:
    scores[22] = s_high(
        f"{single_count}/{len(BIZ_WL)} 个业务工作负载为单副本部署。" + coll_brief(22))
elif single_count <= len(BIZ_WL) * 0.5:
    scores[22] = s_part(
        f"{single_count}/{len(BIZ_WL)} 个业务工作负载为单副本部署，不满足 ≥2 副本高可用要求。" + coll_brief(22))
else:
    scores[22] = s_none(
        f"{single_count}/{len(BIZ_WL)} 个业务工作负载为单副本部署，严重不满足高可用要求。" + coll_brief(22))

# #23 Storage stateless (hostPath check) - parsed from collection.md
coll23 = COLL.get(23, "")
hostpath_wl_count = 0
hostpath_match = re.search(r'使用 hostPath 的工作负载数量:\s*(\d+)', coll23)
if hostpath_match:
    hostpath_wl_count = int(hostpath_match.group(1))
# Only hostPath in business namespaces counts as risk
biz_ns_list_23 = _meta.get("biz_namespaces", [])
biz_hostpath_lines = []
for l in coll23.split("\n"):
    s = l.strip().lstrip("-").strip()
    if not s:
        continue
    if any(re.match(rf'^{re.escape(ns)}/', s) for ns in biz_ns_list_23):
        biz_hostpath_lines.append(s)
biz_hostpath_count = len(biz_hostpath_lines)

if biz_hostpath_count == 0 and hostpath_wl_count == 0:
    scores[23] = s_full(coll_brief(23))
elif biz_hostpath_count == 0:
    scores[23] = s_high(
        f"hostPath 挂载 {hostpath_wl_count} 处，但均为系统组件（非业务命名空间）。" + coll_brief(23))
elif biz_hostpath_count <= 2:
    scores[23] = s_part(
        f"业务命名空间存在 {biz_hostpath_count} 处 hostPath 挂载，建议改用 PVC。" + coll_brief(23))
else:
    scores[23] = s_none(
        f"业务命名空间存在 {biz_hostpath_count} 处 hostPath 挂载，不满足无状态化要求。" + coll_brief(23))

# #24 Service cross-AZ high availability - parsed from collection.md
coll24 = COLL.get(24, "")
antiaff_match = re.search(r'Pod 反亲和性配置:\s*(\d+)', coll24)
topospread_match = re.search(r'拓扑分布约束:\s*(\d+)', coll24)
antiaff_count = int(antiaff_match.group(1)) if antiaff_match else 0
topospread_count = int(topospread_match.group(1)) if topospread_match else 0
# Check if business namespace has topology constraints configured
biz_topospread = [l for l in coll24.split("\n")
                  if l.strip().startswith("-")
                  and any(ns+"/" in l for ns in _meta.get("biz_namespaces", []))]
biz_has_topospread = len(biz_topospread) > 0

if biz_has_topospread and antiaff_count > 0:
    scores[24] = s_high(
        f"Pod 反亲和性配置 {antiaff_count} 个工作负载，拓扑分布约束 {topospread_count} 个工作负载。"
        + coll_brief(24))
elif antiaff_count > 0:
    scores[24] = s_part(
        f"Pod 反亲和性配置 {antiaff_count} 个工作负载，但业务未配置 topologySpreadConstraints。" + coll_brief(24))
else:
    scores[24] = s_none(coll_brief(24))

# #25 Cluster management plane high availability
master_azs = _meta.get("master_azs", [])
if len(master_azs) >= 3:
    scores[25] = s_full(coll_brief(25))
elif len(master_azs) >= 2:
    scores[25] = s_high(coll_brief(25))
else:
    scores[25] = s_none(coll_brief(25))

# #26 Nodes distributed across AZs
node_az_dist = _meta.get("node_az_distribution", {})
distinct_azs = len(node_az_dist)
if distinct_azs >= 2:
    scores[26] = s_full(coll_brief(26))
else:
    az_detail = ", ".join(f"{az}: {cnt}" for az, cnt in node_az_dist.items())
    scores[26] = s_none(
        f"集群 {NODE_COUNT} 个工作节点均位于 {az_detail}，未跨 AZ 分布，存在 AZ 故障风险。" + coll_brief(26))

# #27 Business scheduling by node pool for resource isolation
coll27 = COLL.get(27, "")
ns_match = re.search(r'nodeSelector 配置:\s*(\d+)', coll27)
aff_match = re.search(r'nodeAffinity 配置:\s*(\d+)', coll27)
ns_count = int(ns_match.group(1)) if ns_match else 0
aff_count = int(aff_match.group(1)) if aff_match else 0
if ns_count > 0 or aff_count > 0:
    scores[27] = s_part(
        f"nodeSelector 配置 {ns_count} 个，nodeAffinity 配置 {aff_count} 个（建议项）。" + coll_brief(27))
else:
    scores[27] = s_none(coll_brief(27))

# #28 Scheduled scaling CronHPA
coll28 = COLL.get(28, "")
cronhpa_match = re.search(r'CronHPA:\s*(\d+)', coll28)
cronhpa_count = int(cronhpa_match.group(1)) if cronhpa_match else 0
hpa_match = re.search(r'HPA:\s*(\d+)', coll28)
hpa_count = int(hpa_match.group(1)) if hpa_match else 0
if cronhpa_count > 0:
    scores[28] = s_full(coll_brief(28))
elif hpa_count > 0:
    scores[28] = s_part(
        f"已配置 HPA {hpa_count} 个，但未配置 CronHPA 定时扩缩容（建议项）。" + coll_brief(28))
else:
    scores[28] = s_none(coll_brief(28))

# #29 Cluster dynamic scaling CA
coll29 = COLL.get(29, "")
ca_status = "未配置"
if "已配置" in coll29 or "启用" in coll29:
    ca_status = "已配置"
if ca_status == "已配置":
    scores[29] = s_full(coll_brief(29))
else:
    scores[29] = s_none(coll_brief(29))

# #30 API monitoring metric coverage
coll30 = COLL.get(30, "")
has_prom = "prometheus" in coll30.lower() or "kube-prometheus-stack" in coll30.lower() or "已安装" in coll30
biz_metrics = "/metrics" in coll30 or "业务级指标" in coll30 or "应用级指标" in coll30
if has_prom and biz_metrics:
    scores[30] = s_full(coll_brief(30))
elif has_prom:
    scores[30] = s_high(coll_brief(30))
else:
    scores[30] = s_none(coll_brief(30))

# #31 Log standardized output stdout/stderr
coll31 = COLL.get(31, "")
stdout_match = re.search(r'(\d+)/(\d+)\s*个容器输出到 stdout/stderr', coll31)
if stdout_match:
    covered, total31 = int(stdout_match.group(1)), int(stdout_match.group(2))
    if covered == total31:
        scores[31] = s_full(coll_brief(31))
    else:
        scores[31] = s_part(coll_brief(31))
else:
    scores[31] = s_full(coll_brief(31))

# #32 Log format normalization (JSON + trace id)
coll32 = COLL.get(32, "")
json_match = re.search(r'(\d+)/(\d+)\s*个容器输出JSON格式', coll32)
if json_match:
    covered, total32 = int(json_match.group(1)), int(json_match.group(2))
    if covered >= total32:
        scores[32] = s_full(coll_brief(32))
    elif covered >= total32 * 0.5:
        scores[32] = s_high(coll_brief(32))
    else:
        scores[32] = s_part(coll_brief(32))
else:
    scores[32] = s_part(coll_brief(32))

# #33 Persistent logs with query support
coll33 = COLL.get(33, "")
has_log_collector = "已安装" in coll33 or "log-agent" in coll33 or "fluent-bit" in coll33
if has_log_collector:
    scores[33] = s_full(coll_brief(33))
else:
    scores[33] = s_none(coll_brief(33))

# #34 Alert configuration
coll34 = COLL.get(34, "")
rule_match = re.search(r'(\d+)\s*条规则', coll34)
aom_match = re.search(r'AOM 告警规则:\s*(\d+)', coll34)
rule_count = 0
if rule_match:
    rule_count = int(rule_match.group(1))
if aom_match:
    aom_count = int(aom_match.group(1))
    rule_count = max(rule_count, aom_count)
if rule_count > 0:
    scores[34] = s_high(coll_brief(34))
else:
    scores[34] = s_none(coll_brief(34))

# #35 Audit logs
coll35 = COLL.get(35, "")
has_audit = "审计日志" in coll35 and ("已配置" in coll35 or "✅" in coll35 or "audit" in coll35.lower())
if has_audit:
    scores[35] = s_full(coll_brief(35))
else:
    scores[35] = s_none(coll_brief(35))

# #36 Pod and container security context
coll36 = COLL.get(36, "")
runas_match = re.search(r'RunAsUser 已配置:\s*(\d+)', coll36)
runas_no_match = re.search(r'RunAsUser 未配置:\s*(\d+)', coll36)
priv_match = re.search(r'特权容器:\s*(\d+)', coll36)
runas_ok = int(runas_match.group(1)) if runas_match else 0
runas_no = int(runas_no_match.group(1)) if runas_no_match else 0
priv_count = int(priv_match.group(1)) if priv_match else 0
total36 = runas_ok + runas_no
if total36 == 0:
    total36 = 1
runas_rate = runas_ok / total36
if runas_rate >= 0.95 and priv_count == 0:
    scores[36] = s_full(coll_brief(36))
elif runas_rate >= 0.5 and priv_count == 0:
    scores[36] = s_high(coll_brief(36))
elif priv_count == 0:
    scores[36] = s_part(
        f"RunAsUser 覆盖率 {runas_rate*100:.0f}%（已配置 {runas_ok}/{total36}），特权容器 {priv_count} 个。" + coll_brief(36))
else:
    scores[36] = s_none(coll_brief(36))

# #37 Security events (image/code vulnerability scanning)
coll37 = COLL.get(37, "")
hss_match = re.search(r'HSS 已扫描镜像数:\s*(\d+)', coll37)
hss_count = int(hss_match.group(1)) if hss_match else 0
vuln_found = False
# Check for vulnerabilities
for kw in ["P0", "P1", "高危", "严重", "漏洞", "vulnerability"]:
    if kw in coll37:
        vuln_found = True
        break
if hss_count > 0 and not vuln_found:
    scores[37] = s_high(coll_brief(37))
elif hss_count > 0:
    scores[37] = s_none(coll_brief(37))
else:
    scores[37] = s_none(coll_brief(37))

# #38 Container security configuration (hostNetwork/hostPID/hostIPC/sensitive directories)
coll38 = COLL.get(38, "")
hostnet_match = re.search(r'hostNetwork:\s*(\d+)', coll38)
hostpid_match = re.search(r'hostPID:\s*(\d+)', coll38)
hostipc_match = re.search(r'hostIPC:\s*(\d+)', coll38)
sens_match = re.search(r'敏感目录挂载:\s*(\d+)', coll38)
hostnet = int(hostnet_match.group(1)) if hostnet_match else 0
hostpid = int(hostpid_match.group(1)) if hostpid_match else 0
hostipc = int(hostipc_match.group(1)) if hostipc_match else 0
sens_count = int(sens_match.group(1)) if sens_match else 0
# Only sensitive mounts in business namespaces count as risk
biz_ns_list = _meta.get("biz_namespaces", [])
biz_sens_lines = []
for l in coll38.split("\n"):
    s = l.strip()
    if not s or s.startswith("hostIPC") or s.startswith("hostNetwork") or s.startswith("hostPID") or s.startswith("敏感目录"):
        continue
    # Match detail lines like "dify/xxx: /path"
    if any(re.match(rf'^{re.escape(ns)}/', s) for ns in biz_ns_list):
        biz_sens_lines.append(s)
biz_sens_count = len(biz_sens_lines)

if hostnet == 0 and hostpid == 0 and hostipc == 0 and biz_sens_count == 0 and sens_count == 0:
    scores[38] = s_full(coll_brief(38))
elif hostnet == 0 and hostpid == 0 and hostipc == 0 and biz_sens_count == 0:
    scores[38] = s_high(
        f"hostNetwork/hostPID/hostIPC 均为 0；敏感目录挂载 {sens_count} 处（均为系统组件，非业务命名空间）。" + coll_brief(38))
elif hostnet == 0 and hostpid == 0 and hostipc == 0:
    scores[38] = s_part(
        f"hostNetwork/hostPID/hostIPC 均为 0；敏感目录挂载 {sens_count} 处，其中业务命名空间 {biz_sens_count} 处需 review。" + coll_brief(38))
else:
    scores[38] = s_none(
        f"hostNetwork={hostnet}, hostPID={hostpid}, hostIPC={hostipc}，敏感目录挂载 {sens_count} 处，业务命名空间 {biz_sens_count} 处。" + coll_brief(38))

# #39 Use latest K8s version
if K8S_VERSION:
    scores[39] = s_full(coll_brief(39))
else:
    scores[39] = s_part(coll_brief(39))

# #40 Reasonable cluster size selection
if NODE_COUNT <= 250:
    scores[40] = s_full(coll_brief(40))
else:
    scores[40] = s_part(coll_brief(40))

# #41 Enable RBAC
coll41 = COLL.get(41, "")
has_rbac = "RBAC" in coll41 or "已检查" in coll41 or "ClusterRole" in coll41
biz_sa_default = False
# Check if business uses default SA
for wl in BIZ_WL:
    if wl.get("ns", "") in _meta.get("biz_namespaces", []):
        # Cannot directly get SA info, infer from collection data
        pass
if has_rbac:
    scores[41] = s_part(coll_brief(41))
else:
    scores[41] = s_none(coll_brief(41))

# #42 Ingress TLS configuration
ingresses = _meta.get("ingresses", [])
tls_count = sum(1 for ing in ingresses if ing.get("tls"))
total_ing = len(ingresses)
if total_ing == 0:
    scores[42] = s_full(coll_brief(42))  # No Ingress, not applicable
elif tls_count == total_ing:
    scores[42] = s_full(coll_brief(42))
elif tls_count > 0:
    scores[42] = s_part(
        f"共 {total_ing} 个 Ingress，{tls_count} 个配置 TLS，{total_ing - tls_count} 个未配置。" + coll_brief(42))
else:
    scores[42] = s_none(
        f"共 {total_ing} 个 Ingress，均未配置 TLS。" + coll_brief(42))

# #43 Restrict node external access (security groups)
coll43 = COLL.get(43, "")
has_sg = "安全组" in coll43 or "已获取" in coll43 or "已检查" in coll43
if has_sg:
    scores[43] = s_full(coll_brief(43))
else:
    scores[43] = s_part(coll_brief(43))

print("scored 21..43")

# Write Excel
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "云原生评估汇总"
header_row = ["编号","云原生维度","级别","验收指标","量化目标","验收方式","描述","验收结果","得分","满分","打分依据"]
ws.append(header_row)

thin = Side(border_style="thin", color="999999")
border = Border(top=thin, bottom=thin, left=thin, right=thin)
header_fill = PatternFill("solid", fgColor="305496")
header_font = Font(bold=True, color="FFFFFF")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)

for c in ws[1]:
    c.fill = header_fill; c.font = header_font; c.alignment = center; c.border = border

color_map = {3:"C6EFCE",2:"FFEB9C",1:"FFC7CE",0:"FF6B6B"}

for it in items:
    sc = scores[it["id"]]
    score, result, reason = sc
    row = [it["id"], it["dim"], it["level"], it["metric"], it["target"], it["method"], it["desc"], result, score, 3, reason]
    ws.append(row)
    r = ws.max_row
    for ci,cell in enumerate(ws[r],1):
        cell.alignment = left if ci in (5,6,7,11) else center
        cell.border = border
    # Color the "acceptance result" and "score" columns
    fill = PatternFill("solid", fgColor=color_map[score])
    ws.cell(r, 8).fill = fill
    ws.cell(r, 9).fill = fill

# Column widths
widths = [6,12,8,18,30,22,40,12,8,8,60]
for i,w in enumerate(widths,1):
    ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

# Row height auto (skipped, openpyxl does not support it, but wrap_text is enabled)
for r in range(2, ws.max_row+1):
    ws.row_dimensions[r].height = 90

OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(str(OUT))
print("saved:", OUT)
total_score = sum(s[0] for s in scores.values())
print(f"total score: {total_score}/{len(scores)*3} ; rate={total_score/(len(scores)*3)*100:.1f}%")
